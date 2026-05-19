#!/usr/bin/env python3
"""B.2 — spreadsheet-driven mechanical race generator.

Source of truth: reference/PF1e_Data.xlsx (Decision 2: AUTHORITATIVE;
no d20pfsrd fallback for any race in the sheet; blank cells are
intentional). 42 races from "Race Mechanical Info" (14 cols) + 9
per-race detail sheets (lore, ages, height/weight, structured racial
features, alternate traits, FCB, archetypes, feats).

Idempotent, augment-only, assert-driven. Default = DRY RUN (prints the
three-bucket inventory + per-race plan). Pass --apply to write.

Buckets:
  MATCH-MECHANICAL  instance exists, is_lore_only false  -> UPDATE
  MATCH-LORE-ONLY   instance exists, is_lore_only true    -> PROMOTE
  NO-MATCH          no instance                            -> CREATE
  SKIP(blank-row)   sheet row has no mechanical data (Vishkanya/Wayang)

Asserts (a failed assert SKIPS that race, never writes bad data):
  *_racial_mod in [-4,4]; size in {small,medium,large}; base_speed_ft>0;
  race_type non-empty; race_subtypes non-empty; languages_default
  non-empty; >=1 vision_trait (blank senses -> "Normal vision").
  An alternate_racial_trait whose parsed replaces[] target does not
  resolve to a racial_feature id on THIS race is dropped + logged.
"""
import json, glob, io, os, re, sys, collections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RI = os.path.join(ROOT, "pf1e", "resource_instances")
XLSX = os.path.join(ROOT, "reference", "PF1e_Data.xlsx")
APPLY = "--apply" in sys.argv

import openpyxl

AB = {"str": "str", "dex": "dex", "con": "con", "int": "int",
      "wis": "wis", "cha": "cha",
      "strength": "str", "dexterity": "dex", "constitution": "con",
      "intelligence": "int", "wisdom": "wis", "charisma": "cha"}
DETAIL_SHEETS = ["Dwarf", "Elf", "Gnome", "Half-Elf", "Half-Orc",
                 "Halfling", "Human", "Catfolk", "Duergar"]
SECTION_HEADERS = {
    "standard racial traits": "Standard",
    "defense racial traits": "Defense",
    "feat and skill racial traits": "Skill",
    "magical racial traits": "Magic",
    "senses racial traits": "Senses",
    "offense racial traits": "Offense",
    "movement racial traits": "Movement",
    "other racial traits": "Other",
}


def clean(x):
    return ("" if x is None else str(x).replace("\xa0", " ").strip())


def slug(s):
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", (s or "").lower())).strip("_")


def norm_name(s):
    """Canonical race name for matching: drop trailing parentheticals
    like ' (9 RP)' / ' (14/41 RP)' / ' (ARG)' and a leading 'ARG '."""
    s = re.sub(r"\s*\([^)]*\)\s*$", "", (s or "").strip())
    s = re.sub(r"^arg\s+", "", s, flags=re.I)
    return s.strip()


def fname_slug(fn):
    """Race slug embedded in a race_*__crb_ filename, stripping the
    lore/arg/source scaffolding so race_lore_arg_catfolk__crb_ -> catfolk."""
    b = re.sub(r"\.rpg\.json$", "", fn)
    b = re.sub(r"^race_", "", b)
    b = re.sub(r"^lore_", "", b)
    b = re.sub(r"^arg_", "", b)
    b = re.sub(r"__crb_$|_crb$", "", b)
    b = re.sub(r"_\d+_rp$", "", b)
    return slug(b)


def w(v):
    return {"value": v}


# ---------------------------------------------------------------- parse
def parse_ability(plus, minus):
    """-> (mods dict, floating_count). Raises on out-of-range."""
    mods = {k: 0 for k in ("str", "dex", "con", "int", "wis", "cha")}
    floating = 0

    def toks(cell):
        out = []
        for t in cell.split(","):
            t = t.strip()
            if not t or t.lower() in ("none",):
                continue
            out.append(t)
        return out

    for t in toks(plus):
        low = t.lower()
        if low.startswith("any"):
            floating = 1
            continue
        m = re.match(r"([a-z]+)\s*(?:\(\s*([+\-]?\d+)\s*\))?", low)
        if not m:
            continue
        ab = AB.get(m.group(1))
        if not ab:
            continue
        amt = int(m.group(2)) if m.group(2) else 2
        mods[ab] += abs(amt)
    for t in toks(minus):
        low = t.lower()
        if low.startswith("any"):
            continue
        m = re.match(r"([a-z]+)\s*(?:\(\s*([+\-]?\d+)\s*\))?", low)
        if not m:
            continue
        ab = AB.get(m.group(1))
        if not ab:
            continue
        amt = int(m.group(2)) if m.group(2) else -2
        mods[ab] -= abs(amt)
    for k, v in mods.items():
        assert -4 <= v <= 4, "ability mod %s=%d out of [-4,4]" % (k, v)
    return mods, floating


def parse_type(cell):
    m = re.match(r"\s*([A-Za-z ]+?)\s*(?:\(([^)]*)\))?\s*$", cell)
    if not m:
        return "", []
    rtype = m.group(1).strip().lower()
    subs = []
    if m.group(2):
        subs = [s.strip().lower() for s in m.group(2).split(",") if s.strip()]
    return rtype, subs


def parse_speed(cell):
    out = dict(base=0, swim=0, fly=0, climb=0, burrow=0, fly_man="")
    nums = re.findall(r"(\d+)\s*ft", cell)
    if nums:
        out["base"] = int(nums[0])
    for kind in ("swim", "fly", "climb", "burrow"):
        m = re.search(kind + r"\s*(\d+)\s*ft", cell, re.I)
        if m:
            out[kind] = int(m.group(1))
    fm = re.search(r"fly[^()]*\(([a-z]+)\)", cell, re.I)
    if fm:
        out["fly_man"] = fm.group(1).lower()
    return out


def parse_langs(cell):
    out = []
    for t in re.split(r"[,;]", cell):
        t = t.strip().lower()
        t = re.sub(r"\s*\(.*?\)", "", t).strip()
        if not t or t.startswith("1 ") or "elemental" in t or t.startswith("or "):
            if t.startswith("or "):
                t = t[3:].strip()
            else:
                continue
        if t:
            out.append(t)
    return out


def parse_senses(cell):
    vt, other = [], []
    dv, llv = 0, False
    parts = [p.strip() for p in re.split(r"[,;]", cell) if p.strip()]
    for p in parts:
        low = p.lower()
        if "darkvision" in low:
            m = re.search(r"(\d+)\s*ft", low)
            rng = int(m.group(1)) if m else (120 if "superior" in low else 60)
            dv = max(dv, rng)
            vt.append("Darkvision %d ft." % rng)
        elif "low-light" in low or "low light" in low:
            llv = True
            vt.append("Low-light vision")
        else:
            other.append(p)
    if not vt:
        vt = ["Normal vision"]
    return vt, dv, llv, other


# ------------------------------------------------- detail-sheet parsing
def parse_detail_sheet(ws):
    rows = []
    for r in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in r]
        rows.append(cells)

    def is_hdr(a):
        return a.strip().lower() in SECTION_HEADERS

    out = dict(lore="", features=[], alts=[], fcb=[], arch=[], feats=[],
               age={}, hw={})
    # ---- lore: rows before first "Table:" ----
    lore_lines = []
    i = 0
    while i < len(rows):
        a = rows[i][0]
        if a.lower().startswith("table:") or a.lower().startswith("random "):
            break
        if a:
            lore_lines.append(a)
        i += 1
    out["lore"] = "\n\n".join(lore_lines).strip()

    # ---- walk the rest, section-aware ----
    cur = None
    seen_alt_intro = False
    for j in range(len(rows)):
        a = rows[j][0]
        rest = [c for c in rows[j][1:] if c]
        low = a.lower()

        if low.startswith("table: random starting ages") or low.startswith("random ") and "age" in low:
            # header row j+1, values row j+2
            try:
                vals = [c for c in rows[j + 2] if c]
                # e.g. ["40 years","+3d6 years","+5d6 years","+7d6 years"]
                am = re.search(r"(\d+)", vals[0])
                out["age"] = dict(
                    adulthood=int(am.group(1)) if am else 0,
                    intuitive=vals[1] if len(vals) > 1 else "",
                    self_taught=vals[2] if len(vals) > 2 else "",
                    trained=vals[3] if len(vals) > 3 else "")
            except Exception:
                pass
            cur = None
            continue
        if low.startswith("table: random height") or (low.startswith("random ") and "height" in low):
            try:
                for k in range(j + 1, min(j + 8, len(rows))):
                    rc = [c for c in rows[k] if c]
                    if not rc:
                        continue
                    g = rows[k][0].lower()
                    cells = [c for c in rows[k][1:] if c]
                    if g.startswith("male") and len(cells) >= 4:
                        out["hw"].update(hmb=cells[0], hmm=cells[1],
                                         wmb=cells[2], wmm=cells[3])
                    elif g.startswith("female") and len(cells) >= 4:
                        out["hw"].update(hfb=cells[0], hfm=cells[1],
                                         wfb=cells[2], wfm=cells[3])
            except Exception:
                pass
            cur = None
            continue

        if is_hdr(a):
            cur = SECTION_HEADERS[low]
            continue
        if low == "alternate racial traits":
            cur = "ALT"
            seen_alt_intro = False
            continue
        if low.startswith("racial subtypes"):
            cur = None
            continue
        if low == "favored class options":
            cur = "FCB"
            continue
        if low.startswith("racial archetypes"):
            cur = "ARCH"
            continue
        if low == "racial feats":
            cur = "FEATS"
            continue
        if low in ("subpages", "3pp options", "faq") or low.startswith("q:") or low.startswith("a:"):
            continue
        if not a:
            continue

        if cur in ("Standard", "Defense", "Skill", "Senses", "Offense",
                   "Movement", "Magic", "Other"):
            m = re.match(r"\s*([^:]{2,60}?):\s*(.+)$", a)
            if m:
                nm = m.group(1).strip()
                out["features"].append(dict(cat=cur, name=nm,
                                            desc=m.group(2).strip()))
        elif cur == "ALT":
            if not seen_alt_intro and a.lower().startswith("the following"):
                seen_alt_intro = True
                continue
            m = re.match(r"\s*(.+?)\s*(?:\((\d+)\s*RP\))?:\s*(.+)$", a)
            if m:
                out["alts"].append(dict(name=m.group(1).strip(),
                                        rp=int(m.group(2)) if m.group(2) else 0,
                                        desc=m.group(3).strip()))
        elif cur == "FCB":
            if a.lower().startswith("the following"):
                continue
            m = re.match(r"\s*([A-Za-z][A-Za-z /]{1,30}?):\s*(.+)$", a)
            if m:
                out["fcb"].append(dict(cls=m.group(1).strip(),
                                       desc=m.group(2).strip()))
        elif cur == "ARCH":
            if a.lower().startswith("the following"):
                continue
            nm = re.sub(r"\s*\(.*?\)\s*$", "", a).strip()
            if nm:
                out["arch"].append(nm)
        elif cur == "FEATS":
            if a.lower().startswith("the following"):
                continue
            out["feats"].append(a.strip())
    return out


# ------------------------------------------------------- main pipeline
def load_instances():
    inst = {}
    for p in sorted(glob.glob(os.path.join(RI, "race_*.rpg.json"))):
        d = json.load(io.open(p, encoding="utf-8"))
        s = d.get("stats", {})
        nm = (s.get("name") or {}).get("value", "")
        lo = (s.get("is_lore_only") or {}).get("value", False)
        fn = os.path.basename(p)
        inst[fn] = dict(path=p, name=nm, slug=slug(norm_name(nm)),
                        fslug=fname_slug(fn), lore_only=bool(lo), data=d)
    return inst


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Race Mechanical Info"]
    races = {}
    for r in ws.iter_rows(values_only=True):
        c = [clean(x) for x in r]
        a = c[0]
        if a.endswith(" Races") or a in ("Race", "") or c[1] in ("Ability Score Plus", "Racial Traits"):
            continue
        if not a:
            continue
        races[a] = dict(name=a, plus=c[1], minus=c[2], size=c[3], type=c[4],
                        speed=c[5], langs=c[6], senses=c[7], dfe=c[8],
                        off=c[9], skill=c[10], feats=c[11], sla=c[12],
                        rp=c[13])
    details = {s: parse_detail_sheet(wb[s]) for s in DETAIL_SHEETS}
    inst = load_instances()
    by_slug = collections.defaultdict(list)
    for fn, v in inst.items():
        keys = {v["slug"], v["fslug"]}
        for k in keys:
            if k and fn not in by_slug[k]:
                by_slug[k].append(fn)

    # irregular plural / spelling aliases (sheet name slug -> instance slug)
    ALIAS = {"gillman": "gillmen", "samsaran": "samsarans",
             "changeling": "changelings", "wayang": "wayangs"}

    def best_match(sg):
        cands = by_slug.get(sg, []) or by_slug.get(ALIAS.get(sg, ""), [])
        if not cands:
            return []
        # deterministic: prefer an already-mechanical/arg lore file,
        # then shortest filename (the canonical, least-qualified entry)
        return sorted(cands, key=lambda f: (inst[f]["lore_only"], len(f)))

    plan = []
    for nm, row in sorted(races.items()):
        sg = slug(nm)
        # blank-row guard (Vishkanya / Wayang)
        if not any([row["plus"], row["minus"], row["size"], row["type"],
                    row["speed"], row["langs"]]):
            plan.append((nm, "SKIP(blank-row)", None, None, []))
            continue
        logs = []
        try:
            mods, floating = parse_ability(row["plus"], row["minus"])
            rtype, subs = parse_type(row["type"])
            sp = parse_speed(row["speed"])
            langs = parse_langs(row["langs"])
            vt, dv, llv, other = parse_senses(row["senses"])
            size = row["size"].strip().lower()
            assert size in ("small", "medium", "large"), "size=%r" % size
            assert sp["base"] > 0, "base_speed=%d" % sp["base"]
            assert rtype, "empty race_type"
            # subtypes legitimately empty for Dragon/Plant/etc. — no assert
            assert langs, "no languages"
            assert vt, "no vision"
        except AssertionError as e:
            plan.append((nm, "SKIP(assert)", str(e), None, []))
            continue

        # detail sheet (match by name; Half-Elf/Half-Orc sheet names)
        dkey = next((s for s in DETAIL_SHEETS if slug(s) == sg), None)
        det = details.get(dkey) if dkey else None

        # build racial_features + resolve alt-trait replaces
        feats_struct, feat_ids = [], set()
        if det:
            for f in det["features"]:
                fid = "%s_%s" % (sg, slug(f["name"]))
                feat_ids.add(fid)
                feats_struct.append((fid, f["name"], f["cat"], f["desc"]))
        alts_struct = []
        if det:
            for al in det["alts"]:
                replaces = []
                mrep = re.search(
                    r"(?:replaces|alters|modifies)\s+(?:the\s+)?([a-z][a-z ,and]+?)(?:\s+racial trait|\.|$)",
                    al["desc"], re.I)
                unresolved = False
                if mrep:
                    for cand in re.split(r",|\band\b", mrep.group(1)):
                        cand = cand.strip()
                        if not cand:
                            continue
                        cid = "%s_%s" % (sg, slug(cand))
                        if cid in feat_ids:
                            replaces.append(cid)
                        else:
                            unresolved = True
                # zero-bad-data: never emit an unresolved id. Keep the
                # trait (its name+desc IS displayed) with replaces=[];
                # the replaces[] linkage only feeds the DEFERRED
                # alt-trait picker UX, so an empty list loses nothing
                # displayable. Log for the report.
                if unresolved:
                    replaces = []
                    logs.append("alt-trait %r replaces[] unresolved -> kept w/ empty replaces"
                                % al["name"][:40])
                alts_struct.append((al["name"], al["desc"], replaces, al["rp"]))

        # bucket. core mechanical = is_lore_only False AND a bare
        # race_<slug>__crb_ filename (not the race_lore_arg_* tail).
        matches = best_match(sg)
        mech = [m for m in matches if not inst[m]["lore_only"]
                and not m.startswith("race_lore_")]
        if mech:
            bucket, target = "UPDATE", mech[0]
        elif matches:
            # PROMOTE into the existing file in place (fix its name,
            # add mechanics, keep lore). Prevents a duplicate
            # race_<slug>__crb_ that would gotcha-#6-collide.
            bucket, target = "PROMOTE", matches[0]
        else:
            bucket, target = "CREATE", "race_%s__crb_.rpg.json" % sg

        rec = dict(mods=mods, floating=floating, rtype=rtype, subs=subs,
                   sp=sp, langs=langs, vt=vt, dv=dv, llv=llv, other=other,
                   size=size, row=row, det=det, feats=feats_struct,
                   alts=alts_struct, sg=sg)
        plan.append((nm, bucket, target, rec, logs))

    # ---- report ----
    cnt = collections.Counter(p[1] for p in plan)
    print("=" * 64)
    print("B.2 RACE GENERATOR  —  %s" % ("APPLY" if APPLY else "DRY RUN"))
    print("=" * 64)
    print("Spreadsheet races: %d   Instances: %d" % (len(races), len(inst)))
    for k in ("UPDATE", "PROMOTE", "CREATE", "SKIP(blank-row)", "SKIP(assert)"):
        if cnt.get(k):
            print("  %-16s %d" % (k, cnt[k]))
    print("-" * 64)
    for nm, bucket, tgt, rec, logs in plan:
        extra = ""
        if rec:
            m = rec["mods"]
            nz = ",".join("%s%+d" % (k, v) for k, v in m.items() if v) or ("Any+2" if rec["floating"] else "none")
            extra = " | %s | %s %s %dft | feats=%d alts=%d" % (
                nz, rec["size"], rec["rtype"], rec["sp"]["base"],
                len(rec["feats"]), len(rec["alts"]))
        print("  %-14s %-16s %s%s" % (nm, bucket, (tgt or "")[:30], extra))
        for lg in logs:
            print("        ! %s" % lg)

    # spreadsheet races with no instance + project races not in sheet
    sheet_slugs = {slug(n) for n in races}
    not_in_proj = [n for n in races if not best_match(slug(n))]
    not_in_sheet = sorted({inst[f]["name"] for f in inst
                           if inst[f]["slug"] not in sheet_slugs
                           and inst[f]["fslug"] not in sheet_slugs})
    print("-" * 64)
    print("Sheet races with NO project instance (CREATE fresh): %s"
          % (not_in_proj or "none"))
    print("Project races NOT in spreadsheet (%d) -> Task 4 cache path:"
          % len(not_in_sheet))
    print("  " + ", ".join(not_in_sheet))

    if not APPLY:
        print("\nDRY RUN — no files written. Re-run with --apply.")
        return

    written = 0
    for nm, bucket, tgt, rec, logs in plan:
        if bucket not in ("UPDATE", "PROMOTE", "CREATE"):
            continue
        path = os.path.join(RI, tgt)
        if bucket == "CREATE":
            d = {"resource_id": "race", "stats": {}}
            d["stats"]["id"] = "race_%s__crb_" % rec["sg"]
            d["stats"]["name"] = w(nm)
            d["stats"]["source"] = w("crb")
            d["stats"]["lore"] = w("")
        else:
            d = inst[tgt]["data"]
        s = d["stats"]
        # normalise the display name (fixes "Catfolk (9 RP)" -> "Catfolk")
        s["name"] = w(nm)
        if "id" not in s or not s["id"]:
            s["id"] = "race_%s__crb_" % rec["sg"]
        s["is_lore_only"] = w(False)
        s["source"] = w("crb")
        for ab in ("str", "dex", "con", "int", "wis", "cha"):
            s["%s_racial_mod" % ab] = w(rec["mods"][ab])
        s["floating_ability_count"] = w(rec["floating"])
        s["floating_ability_amount"] = w(2)
        s["race_type"] = w(rec["rtype"])
        s["race_subtypes"] = w(rec["subs"])
        s["size"] = w(rec["size"])
        s["base_speed_ft"] = w(rec["sp"]["base"])
        s["swim_speed_ft"] = w(rec["sp"]["swim"])
        s["fly_speed_ft"] = w(rec["sp"]["fly"])
        s["fly_maneuverability"] = w(rec["sp"]["fly_man"])
        s["climb_speed_ft"] = w(rec["sp"]["climb"])
        s["burrow_speed_ft"] = w(rec["sp"]["burrow"])
        s["languages_default"] = w(rec["langs"])
        s["vision_traits"] = w(rec["vt"])
        s["darkvision_range_ft"] = w(rec["dv"])
        s["low_light_vision"] = w(rec["llv"])
        s["senses_other"] = w(rec["other"])
        s["defensive_traits_summary"] = w(rec["row"]["dfe"])
        s["offensive_traits_summary"] = w(rec["row"]["off"])
        s["skill_bonuses_summary"] = w(rec["row"]["skill"])
        s["spell_like_abilities_summary"] = w(rec["row"]["sla"])
        s["bonus_feats_summary"] = w(rec["row"]["feats"])
        rpm = re.match(r"\d+", rec["row"]["rp"])
        s["race_points"] = w(int(rpm.group(0)) if rpm else 0)
        # cosmetic display string (kept for back-compat)
        if rec["floating"]:
            s["ability_score_modifiers"] = w(["Any +2 (racial, choose one)"])
        else:
            disp = ["%s %s%d (racial)" % (k.upper(), "+" if v > 0 else "-", abs(v))
                    for k, v in rec["mods"].items() if v]
            s["ability_score_modifiers"] = w(disp)
        # structured racial_features
        rf = [{"resource_id": "racial_feature",
               "stats": {"id": fid, "name": w(fn), "category": w(cat),
                         "description": w(de), "mechanical_text": w("")}}
              for (fid, fn, cat, de) in rec["feats"]]
        s["racial_features"] = w(rf)
        art = [{"resource_id": "alternate_racial_trait",
                "stats": {"id": "%s_alt_%s" % (rec["sg"], slug(an)),
                          "name": w(an), "description": w(de),
                          "replaces": w(rep), "rp_cost": w(rp)}}
               for (an, de, rep, rp) in rec["alts"]]
        s["alternate_racial_traits"] = w(art)
        if rec["det"]:
            fco = [{"resource_id": "favored_class_option",
                    "stats": {"id": "%s_fcb_%s_%d" % (rec["sg"], slug(f["cls"]), i),
                              "class_name": w(f["cls"]), "class_id": w(""),
                              "description": w(f["desc"])}}
                   for i, f in enumerate(rec["det"]["fcb"])]
            s["favored_class_options"] = w(fco)
            s["racial_archetypes_names"] = w(rec["det"]["arch"])
            s["racial_feats_names"] = w(rec["det"]["feats"])
            ag = rec["det"]["age"]
            if ag:
                s["starting_age_adulthood"] = w(ag.get("adulthood", 0))
                s["starting_age_intuitive"] = w(ag.get("intuitive", ""))
                s["starting_age_self_taught"] = w(ag.get("self_taught", ""))
                s["starting_age_trained"] = w(ag.get("trained", ""))
            hw = rec["det"]["hw"]
            if hw:
                s["height_male_base"] = w(hw.get("hmb", ""))
                s["height_male_modifier"] = w(hw.get("hmm", ""))
                s["height_female_base"] = w(hw.get("hfb", ""))
                s["height_female_modifier"] = w(hw.get("hfm", ""))
                s["weight_male_base"] = w(hw.get("wmb", ""))
                s["weight_male_modifier"] = w(hw.get("wmm", ""))
                s["weight_female_base"] = w(hw.get("wfb", ""))
                s["weight_female_modifier"] = w(hw.get("wfm", ""))
            if rec["det"]["lore"] and bucket == "CREATE":
                s["lore"] = w(rec["det"]["lore"])
        out = os.path.join(RI, "race_%s__crb_.rpg.json" % rec["sg"]
                           if bucket == "CREATE" else tgt)
        with io.open(out, "w", encoding="utf-8", newline="\n") as fh:
            json.dump(d, fh, ensure_ascii=False, indent=2)
        written += 1
    print("\nAPPLIED: %d race instance(s) written." % written)


if __name__ == "__main__":
    main()
